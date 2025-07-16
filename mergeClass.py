
import os
import sys
from typing import List, Dict, Set, Any, Optional
from openpyxl import load_workbook                          # see https://github.com/eea/odfpy
from odf.opendocument import OpenDocumentSpreadsheet, load
from odf.table import Table, TableRow, TableCell
from odf.text import P
from odf.style import Style, TableColumnProperties, TextProperties, TableCellProperties
from odf.number import NumberStyle, Number


default_ODS_name = "EOB_Claims_data.ods"

class ClaimsDataMerger:
    """Handles merging of medical claims data into ODS format."""

    def __init__(self, ods_filename: str = "default_ODS_name"):
        self.ods_filename = ods_filename
        self.expected_headers = [
            "Claim number", "Date of Service", "Date Received", "Date processed",
            "Claim status", "Provider name", "Member name", "Amount billed",
            "Your discounted rate", "Amount we paid", "Amount you may owe",
            "Applies to my deductible", "Copay", "Coinsurance", "Other Insurance",
            "Medication name", "Prescription number", "NDC number", "Day supply",
            "Quantity", "Pharmacy name", "Pharmacy number"
        ]
        self.streamlined_headers = [
            "Claim number", "D.O.S", "Received", "Processed",
            "Status", "Provider", "Patient", "Billed",
            "Your Rate", "We Paid", "You owe",
            "Deductible", "Copay", "Coins.", "Other Ins.",
            "Medication", "Prescription", "NDC number", "Day supply",
            "Quantity", "Pharmacy", "Pharm. number"
        ]

    def read_excel_claims(self, excel_filename: str) -> List[Dict[str, Any]]:
        """Read claims data from Excel file."""
        if not os.path.exists(excel_filename):
            raise FileNotFoundError(f"Excel file '{excel_filename}' not found.")

        try:
            # Try with data_only=True to ignore formulas and formatting
            print(f"Attempting to read Excel file: {excel_filename}")
            workbook = load_workbook(excel_filename, read_only=True, data_only=True)
            sheet = workbook.active

            # Check if headers are in row 1 or row 2
            first_row_headers = []
            for cell in sheet[1]:
                if cell.value:
                    first_row_headers.append(str(cell.value).strip())
                else:
                    first_row_headers.append("")

            second_row_headers = []
            if sheet.max_row >= 2:
                for cell in sheet[2]:
                    if cell.value:
                        second_row_headers.append(str(cell.value).strip())
                    else:
                        second_row_headers.append("")

            # Determine which row contains the actual headers
            if self._validate_headers(second_row_headers):
                headers = second_row_headers
                header_row = 2
                data_start_row = 3
                print(f"Found headers in row 2: {headers}")
            elif self._validate_headers(first_row_headers):
                headers = first_row_headers
                header_row = 1
                data_start_row = 2
                print(f"Found headers in row 1: {headers}")
            else:
                # Try to use row 2 anyway if it looks like headers
                if any(h.lower() in ['claim number', 'date of service'] for h in second_row_headers):
                    headers = second_row_headers
                    header_row = 2
                    data_start_row = 3
                    print(f"Using row 2 as headers (partial match): {headers}")
                else:
                    headers = first_row_headers
                    header_row = 1
                    data_start_row = 2
                    print(f"Using row 1 as headers (default): {headers}")
            #
            # print(f"DEBUG: Using header row {header_row}, data starts at row {data_start_row}")
            # print(f"DEBUG: Found headers: {headers}")

            # Read data rows
            claims = []
            for row_num, row in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
                if any(cell is not None for cell in row):  # Skip empty rows
                    claim_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):   # fill in the claim_dict for this row
                            claim_dict[header] = row[i] if row[i] is not None else ""
                        else:
                            claim_dict[header] = ""
                    Valid = True
                    for k in claim_dict.keys():   # drop "Grand Total:" rows
                        if "Grand" in claim_dict[k]:
                            Valid=False
                    if Valid:
                        claims.append(claim_dict)
                    else:
                        print('Rejecting row: ', row)
                        Valid = False
                    # Debug first few rows
                    if row_num <= data_start_row + 2:
                        print(f"DEBUG: Row {row_num} data: {list(row)[:5]}...")
                        print(f"DEBUG: Claim number for row {row_num}: '{claim_dict.get('Claim number', 'NOT_FOUND')}'")

            workbook.close()
            print(f"DEBUG: Total claims read: {len(claims)}")
            return claims

        except Exception as e:
            print(f"First attempt failed: {str(e)}")

            print("    quitting()")
            quit()

            # print("Trying alternative approach without read_only mode...")
            #
            # # Try without read_only mode as a fallback
            # try:
            #     workbook = load_workbook(excel_filename, data_only=True)
            #     sheet = workbook.active
            #     print("Successfully opened file without read_only mode")
            #
            #     # Get headers
            #     first_row_headers = []
            #     for cell in sheet[1]:
            #         if cell.value:
            #             first_row_headers.append(str(cell.value).strip())
            #         else:
            #             first_row_headers.append("")
            #
            #     second_row_headers = []
            #     if sheet.max_row >= 2:
            #         for cell in sheet[2]:
            #             if cell.value:
            #                 second_row_headers.append(str(cell.value).strip())
            #             else:
            #                 second_row_headers.append("")
            #
            #     if self._validate_headers(second_row_headers):
            #         headers = second_row_headers
            #         data_start_row = 3
            #     elif self._validate_headers(first_row_headers):
            #         headers = first_row_headers
            #         data_start_row = 2
            #     else:
            #         if any(h.lower() in ['claim number', 'date of service'] for h in second_row_headers):
            #             headers = second_row_headers
            #             data_start_row = 3
            #         else:
            #             headers = first_row_headers
            #             data_start_row = 2
            #
            #     print(f"Fallback headers: {headers}")
            #
            #     claims = []
            #     for row_num in range(data_start_row, sheet.max_row + 1):
            #         row_data = []
            #         for col_num in range(1, len(headers) + 1):
            #             cell_value = sheet.cell(row=row_num, column=col_num).value
            #             row_data.append(cell_value if cell_value is not None else "")
            #
            #         if any(cell for cell in row_data):
            #             claim_dict = {}
            #             for i, header in enumerate(headers):
            #                 if i < len(row_data):
            #                     claim_dict[header] = row_data[i]
            #                 else:
            #                     claim_dict[header] = ""
            #             claims.append(claim_dict)
            #
            #     workbook.close()
            #     print(f"Successfully read {len(claims)} claims using fallback method")
            #     return claims
            #
            # except Exception as e2:
            #     raise Exception(f"Error reading Excel file '{excel_filename}': {str(e)}. Fallback also failed: {str(e2)}")

    def _validate_headers(self, headers: List[str]) -> bool:
        """Validate that headers match expected format."""
        # Clean headers for comparison (remove extra spaces, case insensitive)
        clean_headers = [h.strip().lower() for h in headers if h.strip()]
        clean_expected = [h.strip().lower() for h in self.expected_headers]

        return clean_headers == clean_expected

    def load_or_create_ods(self) -> OpenDocumentSpreadsheet:
        """Load existing ODS file or create new one."""
        if os.path.exists(self.ods_filename):
            try:
                doc = load(self.ods_filename)
                print(f"Loaded existing ODS file: {self.ods_filename}")
                return doc
            except Exception as e:
                print(f"Error loading ODS file: {str(e)}")
                print("Creating new ODS file...")
                return self._create_new_ods()
        else:
            print(f"Creating new ODS file: {self.ods_filename}")
            return self._create_new_ods()

    def _create_new_ods(self) -> OpenDocumentSpreadsheet:
        """Create new ODS document with headers."""
        doc = OpenDocumentSpreadsheet()

        # Create main table
        table = Table(name="Claims")

        # Add header row
        header_row = TableRow()
        for header in self.streamlined_headers:
            cell = TableCell()
            cell.addElement(P(text=header))
            header_row.addElement(cell)
        table.addElement(header_row)

        doc.spreadsheet.addElement(table)
        return doc

    def get_existing_claim_numbers(self, doc: OpenDocumentSpreadsheet) -> Set[str]:
        """Get set of existing claim numbers to avoid duplicates."""
        existing_claims = set()

        # Get the first table
        tables = doc.spreadsheet.getElementsByType(Table)
        if not tables:
            print("No existing tables found.")
            return existing_claims

        table = tables[0]
        rows = table.getElementsByType(TableRow)

        if len(rows) < 2:  # No data rows
            print("No existing data rows found.")
            return existing_claims

        # Find claim number column index
        header_row = rows[0]
        header_cells = header_row.getElementsByType(TableCell)

        claim_col_index = 0

        # for i, cell in enumerate(header_cells):
        #     cell_contents = cell.getElementsByType(P)
        #     if cell_contents:
        #         cell_text = str(cell_contents[0]).strip()
        #         if cell_text.lower() == "claim number":
        #             claim_col_index = i
        #             break
        #
        # if claim_col_index == -1:
        #     print("Warning: Could not find 'Claim number' column in existing data.")
        #     return existing_claims

        # Extract existing claim numbers
        for row in rows:  # Skip header row
            cells = row.getElementsByType(TableCell)
            if str(cells[0]).strip() != '' and str(cells[0]).strip().lower() != 'claim number':  # skip blank lines and header
                if len(cells) > claim_col_index:
                    cell_contents = cells[claim_col_index].getElementsByType(P)
                    if cell_contents:
                        claim_num = str(cell_contents[0])
                        if claim_num:
                            existing_claims.add(claim_num)
            else:
                print('Im dropping a header row: ', cells)

        print(f"Found {len(existing_claims)} existing claims.")
        return existing_claims

    def add_claims_to_ods(self, doc: OpenDocumentSpreadsheet, new_claims: List[Dict[str, Any]]) -> int:
        """Add new claims to ODS document, avoiding duplicates."""
        existing_claim_nums = self.get_existing_claim_numbers(doc)
        print(f"Found {len(existing_claim_nums)} existing claims in ODS file.")

        # Get the original ods table
        tables = doc.spreadsheet.getElementsByType(Table)
        if not tables:
            print("Error: No table found in ODS document.")
            return 0

        table = tables[0]
        added_count = 0

        for i, claim in enumerate(new_claims):
            claim_number = str(claim.get("Claim number", "")).strip()
            print(f"Processing claim {i+1}/{len(new_claims)}: {claim_number}")

            # Skip if claim already exists
            if claim_number in existing_claim_nums:
                print(f"  Skipping duplicate claim: {claim_number}")
                continue

            # Create new row
            new_row = TableRow()

            for header in self.expected_headers:
                cell = TableCell()
                value = claim.get(header, "")  # get cell data from dict indexed by header
                if value is None:
                    value = ""

                # Convert value to string and handle special cases
                str_value = str(value).strip()
                tst_value = str_value
                # Add value type attribute for better compatibility
                for c in '.,-$':
                    tst_value = tst_value.replace(c,'')
                if tst_value.isdigit():
                    print('isdigit! ',tst_value)
                    # Numeric value

                    try:
                        # <table:table-cell office:value-type="currency" office:currency="USD" office:value="100.12" calcext:value-type="currency"><text:p>$100.12</text:p></table:table-cell>
                        # cell.setAttribute('office:value-type', 'currency')
                        # cell.setAttribute('office:currency', 'USD')
                        # cell.setAttribute('calcext:value-type','currency')
                        cell.setAttribute('value',P(text=str_value))
                    except ValueError:
                        print("Unknown error line 324")
                        cell.setAttribute('valuetype', 'string')
                else:
                    # String value
                    print('string! ',tst_value)
                    cell.setAttribute('valuetype', 'string')

                cell.addElement(P(text=str_value))
                new_row.addElement(cell)

            table.addElement(new_row)
            existing_claim_nums.add(claim_number)
            added_count += 1
            print(f"  Added claim: {claim_number}")

        print(f"Total claims processed: {len(new_claims)}")
        print(f"Claims added to ODS: {added_count}")
        return added_count

    def save_ods(self, doc: OpenDocumentSpreadsheet):
        """Save ODS document."""
        try:
            doc.save(self.ods_filename)
            print(f"Successfully saved: {self.ods_filename}")

            # Verify the save by reading back
            self.verify_ods_content()

        except Exception as e:
            raise Exception(f"Error saving ODS file: {str(e)}")

    # def clean_ods(self, doc: OpenDocumentSpreadsheet):
    #
    #     print("Cleaning up entries")
    #     tables = doc.spreadsheet.getElementsByType(Table)  #(I think "tables" are "Tabs")
    #
    #     if not tables:
    #         print("Error: No tables found in ODS document.")
    #         return False
    #
    #     patientNameCol = 6  # we are going to shorten patient name
    #     table = tables[0]
    #     rows = table.getElementsByType(TableRow)
    #     for row_idx, row in enumerate(rows):  # Skip header
    #             cells = row.getElementsByType(TableCell)
    #             # eliminate blank lines and header
    #             if str(cells[0]).strip() != '' and str(cells[0]).strip().lower() != 'claim number':
    #                 nametxt = str(cells[patientNameCol]  )
    #                 nametxt = nametxt.replace('Ruggeiro','').replace('Hannaford','') # delete last names
    #                 cells[patientNameCol].value=None  # TODO:  This doesn't work and nobody knows how to replace (not add to) text!!!!!
    #                 cells[patientNameCol].addElement(P(text=nametxt) )
    #     return doc





#  NEW  sort ods rows by date of service
    def sort_ods_by_date_of_service(self, doc: OpenDocumentSpreadsheet) -> bool:
        """Sort the ODS document by Date of Service column (mm/dd/yyyy format), most recent first."""
        if True:  #try
            from datetime import datetime

            print("Sorting document by Date of Service...")
            tables = doc.spreadsheet.getElementsByType(Table)  #(I think "tables" are "Tabs")

            if not tables:
                print("Error: No tables found in ODS document.")
                return False

            table = tables[0]
            rows = table.getElementsByType(TableRow)

            if len(rows) < 2:
                print("No data rows to sort.")
                return True

            # Find the Date of Service column index
            date_col_index = 1
            # header_row = rows[0]
            # header_cells = header_row.getElementsByType(TableCell)
            # date_col_index = -1
            #
            # for i, cell in enumerate(header_cells):
            #     cell_contents = cell.getElementsByType(P)
            #     if cell_contents:
            #         cell_text = str(cell_contents[0]).strip()
            #         if cell_text.lower() == "date of service":
            #             date_col_index = i
            #             break
            #
            # if date_col_index == -1:
            #     print("Error: Could not find 'Date of Service' column.")
            #     return False
            #
            # print(f"Found 'Date of Service' column at index {date_col_index}")

            # Extract data rows with their date values for sorting
            data_rows = []
            for row_idx, row in enumerate(rows):  # Skip header
                cells = row.getElementsByType(TableCell)

                # eliminate blank lines and header
                if str(cells[0]).strip() != '' and str(cells[0]).strip().lower() != 'claim number':

                    # Extract all cell values for this row
                    row_data = []
                    for cell in cells:
                        cell_contents = cell.getElementsByType(P)
                        if cell_contents:
                            row_data.append(str(cell_contents[0]) )
                        else:
                            row_data.append("")

                    # Skip completely blank rows
                    if not any(cell.strip() for cell in row_data):
                        continue

                    # Parse the date from the Date of Service column
                    date_str = row_data[date_col_index] if date_col_index < len(row_data) else ""

                    try:
                        # Parse mm/dd/yyyy format (American format)
                        parsed_date = datetime.strptime(date_str.strip(), "%m/%d/%Y")
                    except ValueError as e:
                        # If parsing fails, use a default old date to put it at the end
                        parsed_date = datetime(1900, 1, 1)
                        print(f"Warning: Could not parse date '{date_str}' in row {row_idx} - Error: {e}")

                    # Store the actual row object with its date for sorting
                    data_rows.append((parsed_date, row))

            # Sort by date (reverse=True for most recent first)
            data_rows.sort(key=lambda x: x[0], reverse=True)
            print(f"Sorted {len(data_rows)} data rows by Date of Service (most recent first)")

            # Remove all  rows from the table
            data_rows_to_remove = rows[:]  # All
            for row in data_rows_to_remove:
                table.removeChild(row)

            # # set up boldface for header
            # EOBboldstyle = Style(name="EOBBoldStyle", family="table-cell")
            # EOBboldstyle.addElement(TextProperties(attributes={"fo:font-weight": "bold"}))
            # doc.styles.addElement(EOBboldstyle)  # put the style into the doc


            ####
            # Add blank lines and header
            row1 = TableRow()
            row2 = TableRow()
            c1 = TableCell(valuetype='string')
            c1.addElement(P(text=''))  # a blank text cell

            for i in range(len(self.streamlined_headers)):
                row1.addElement(c1) # blank string cell
                c2 = TableCell()
                c2.setAttribute('valuetype', 'string')
                c2.addElement(P(text=self.streamlined_headers[i]) )
                row2.addElement(c2) # header row

            table.addElement(row1) # blank line
            table.addElement(row2) # streamlined_headers

            ####
            # Add sorted rows back to table (preserving original formatting)
            for parsed_date, row in data_rows:
                table.addElement(row)

            print("Successfully sorted document in memory")
            return True

        # except Exception as e:
        #     print(f"Error sorting ODS document: {str(e)}")
        #     print(' ... quitting now')
        #     quit()
        #     return False

    def verify_ods_content(self):
        """Verify the content of the saved ODS file."""
        try:
            print("Verifying ODS file content...")
            doc = load(self.ods_filename)
            tables = doc.spreadsheet.getElementsByType(Table)

            if not tables:
                print("ERROR: No tables found in saved file!")
                return

            table = tables[0]
            rows = table.getElementsByType(TableRow)
            print(f"Found {len(rows)} rows in saved file (including header).")

            # Show first few rows for debugging
            for i, row in enumerate(rows[:3]):  # Show first 3 rows
                cells = row.getElementsByType(TableCell)
                row_data = []
                for cell in cells:
                    cell_contents = cell.getElementsByType(P)
                    if cell_contents:
                        row_data.append(str(cell_contents[0]).strip())
                    else:
                        row_data.append("")
                print(f"Row {i+1}: {row_data[:3]}...")  # Show first 3 columns

        except Exception as e:
            print(f"Error verifying ODS content: {str(e)}")

    def merge_excel_to_ods(self, excel_filename: str) -> int:
        """Main method to merge Excel claims data into ODS file."""
        print(f"Reading claims from: {excel_filename}")
        claims = self.read_excel_claims(excel_filename)
        print(f"Found {len(claims)} claims in Excel file.")

        # Debug: Show sample of what was read
        if claims:
            print(f"DEBUG: First claim data: {claims[0]}")
            print(f"DEBUG: Keys in first claim: {list(claims[0].keys())}")
            print(f"DEBUG: Claim number from first claim: '{claims[0].get('Claim number', 'KEY_NOT_FOUND')}'")
        else:
            print("DEBUG: No claims data was read!")

        print(f"Loading ODS file: {self.ods_filename}")
        doc = self.load_or_create_ods()

        print("Merging claims data...")
        added_count = self.add_claims_to_ods(doc, claims)
        print("Sorting...")
        self.sort_ods_by_date_of_service(doc)
        if added_count > 0:
            # self.save_ods(self.clean_ods(doc))
            self.save_ods(doc)
            print(f"Added {added_count} new claims to {self.ods_filename}")
        else:
            print("No new claims to add (all claims already exist).")

        return added_count

